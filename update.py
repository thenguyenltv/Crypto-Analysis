import requests
import json
from openpyxl import load_workbook

FIAT = 'USD'

coins = []
prices = []

def get_all_coin_info():    
    url = 'https://api.coingecko.com/api/v3/coins/markets?vs_currency=usd&order=market_cap_desc&per_page=22000&page=1&sparkline=false'
    response = requests.get(url)
    data = json.loads(response.text)
    for coin_info in data:
        coin = coin_info['symbol'].upper()
        price = coin_info['current_price']
        coins.append(coin)
        prices.append(price)

def get_coin_info(coin):
    url = f'https://min-api.cryptocompare.com/data/pricemultifull?fsyms={coin}&tsyms={FIAT}'
    response = requests.get(url)
    data = json.loads(response.text)
    if 'RAW' not in data:
        return f'Coin {coin} not found!!'
    price = data['RAW'][coin][FIAT]['PRICE']
    coins.append(coin)
    prices.append(price)
    return price

def get_price(coin): 
    coin = coin.upper().strip()       
    for i in range(len(coins)):
        if coins[i] == coin:
            return prices[i]
    return get_coin_info(coin)
            
get_all_coin_info()

file = 'Crypto.xlsx'
wb = load_workbook(filename = file)
sheet = wb['Analysis']

for i in range(3, 3 + 10000):
    coin = sheet.cell(row=i, column=3).value
    if coin == None:
        break
    sheet.cell(row=i, column=5).value = get_price(coin)
wb.save(filename = file)
print('Prices updated successfully!')
print('if you like this project, please give it a star on github: github.com/thenguyenltv/Crypto-Analysis')

